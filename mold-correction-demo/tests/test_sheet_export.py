"""회사 원본 없이 보정시트 엑셀 생성 계약을 검증한다."""

from __future__ import annotations

import re
import sys
import tempfile
import unittest
import zipfile
from pathlib import Path

import numpy as np
import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT))

from sheet_export import (  # noqa: E402
    SheetPoint, SheetView, TitleBlock, build_sheet, crop_view, default_layout,
    place_labels,
)
from sheet_export import config, drawing  # noqa: E402


def _image(width: int = 200, height: int = 120) -> np.ndarray:
    image = np.full((height, width, 3), 255, dtype=np.uint8)
    image[10:-10, 10:-10] = (60, 200, 60)
    return image


def _points(count: int = 4) -> list[SheetPoint]:
    spots = [(0.1, 0.2), (0.9, 0.3), (0.5, 0.05), (0.5, 0.95), (0.25, 0.6)]
    return [
        SheetPoint(f"P-{index:02d}", f"{index * 0.1:+.1f}", *spots[index % len(spots)])
        for index in range(count)
    ]


def _drawing_xml(path: Path) -> str:
    with zipfile.ZipFile(path) as archive:
        name = next(n for n in archive.namelist() if drawing.DRAWING_PART.match(n))
        return archive.read(name).decode("utf-8")


class DrawingXmlTest(unittest.TestCase):
    def test_shapes_use_the_default_namespace(self) -> None:
        # openpyxl 은 spreadsheetDrawing 을 기본 네임스페이스로 쓴다. xdr: 접두사를
        # 붙이면 Excel 이 도형을 통째로 무시한다.
        box = drawing.text_box(1, "-0.7", 10, 20)
        line = drawing.leader(2, 0, 0, 30, 40)

        for xml in (box, line):
            self.assertNotIn("xdr:", xml)
            self.assertTrue(xml.startswith("<absoluteAnchor>"))

    def test_leader_direction_uses_flips_not_negative_size(self) -> None:
        backwards = drawing.leader(1, 100, 100, 20, 20)

        self.assertIn('flipH="1"', backwards)
        self.assertIn('flipV="1"', backwards)
        self.assertNotIn("-", re.search(r'<ext cx="(-?\d+)" cy="(-?\d+)"', backwards).group(0))

    def test_zero_length_leader_still_has_positive_extent(self) -> None:
        line = drawing.leader(1, 50, 50, 50, 50)

        extent = re.search(r'<ext cx="(\d+)" cy="(\d+)"/>', line)
        self.assertGreater(int(extent.group(1)), 0)
        self.assertGreater(int(extent.group(2)), 0)

    def test_label_text_is_escaped(self) -> None:
        xml = drawing.text_box(1, "A & B <2>", 0, 0)

        self.assertIn("A &amp; B &lt;2&gt;", xml)

    def test_inject_requires_a_closing_tag(self) -> None:
        with self.assertRaises(ValueError):
            drawing.inject("<wsDr>", ["<absoluteAnchor/>"])


class LayoutTest(unittest.TestCase):
    def test_default_layout_keeps_the_front_view_below_the_title_block(self) -> None:
        view = SheetView(image=_image(), points=_points())

        default_layout([view])

        x, y, width, height = view.box
        self.assertGreaterEqual(y, config.DRAWING_TOP)
        self.assertLessEqual(y + height, config.DRAWING_BOTTOM)
        self.assertLessEqual(x + width, config.SHEET_WIDTH)

    def test_details_are_laid_out_below_the_front_view(self) -> None:
        front = SheetView(image=_image(), points=_points())
        details = [SheetView(image=_image(80, 80)) for _ in range(2)]

        default_layout([front, *details])

        self.assertLess(front.box[1] + front.box[3], details[0].box[1])
        self.assertLess(details[0].box[0], details[1].box[0])

    def test_caller_supplied_box_is_left_alone(self) -> None:
        view = SheetView(image=_image(), points=[], box=(1.0, 2.0, 3.0, 4.0))

        default_layout([view])

        self.assertEqual(view.box, (1.0, 2.0, 3.0, 4.0))

    def test_fit_box_preserves_aspect(self) -> None:
        view = SheetView(image=_image(200, 100))
        default_layout([view])

        _, _, width, height = view.box

        self.assertAlmostEqual(width / height, 2.0, places=3)

    def test_labels_sit_outside_the_image(self) -> None:
        view = SheetView(image=_image(), points=_points(5))
        default_layout([view])
        box_x, box_y, box_width, box_height = view.box

        for placed in place_labels(view):
            outside = (
                placed.label_x + config.LABEL_WIDTH <= box_x + 1
                or placed.label_x >= box_x + box_width - 1
                or placed.label_y + config.LABEL_HEIGHT <= box_y + 1
                or placed.label_y >= box_y + box_height - 1
            )
            self.assertTrue(outside, f"{placed.text} 라벨이 이미지 위에 있습니다")

    def test_labels_on_one_edge_do_not_overlap(self) -> None:
        # 같은 자리에 모인 포인트들도 서로 밀려나야 한다.
        points = [SheetPoint(f"P-{i}", f"{i}", 0.5, 0.02) for i in range(6)]
        view = SheetView(image=_image(), points=points)
        default_layout([view])

        top = sorted(
            (p for p in place_labels(view) if p.edge == "top"),
            key=lambda p: p.label_x,
        )
        for first, second in zip(top, top[1:]):
            self.assertGreaterEqual(
                second.label_x - first.label_x, config.LABEL_WIDTH
            )

    def test_place_labels_needs_a_box(self) -> None:
        with self.assertRaises(ValueError):
            place_labels(SheetView(image=_image(), points=_points()))

    def test_crop_view_keeps_only_points_inside_and_remaps_them(self) -> None:
        points = [
            SheetPoint("in", "+0.1", 0.60, 0.50),
            SheetPoint("out", "+0.2", 0.10, 0.50),
        ]

        view = crop_view(_image(200, 120), points, (0.5, 0.25, 0.5, 0.5), "DETAIL A")

        self.assertEqual([p.point_id for p in view.points], ["in"])
        self.assertAlmostEqual(view.points[0].x_ratio, 0.2, places=6)
        self.assertAlmostEqual(view.points[0].y_ratio, 0.5, places=6)
        self.assertEqual(view.image.shape[:2], (60, 100))
        self.assertEqual(view.title, "DETAIL A")

    def test_crop_view_rejects_a_degenerate_region(self) -> None:
        with self.assertRaises(ValueError):
            crop_view(_image(), [], (0.5, 0.5, 0.0, 0.5))


class BuildSheetTest(unittest.TestCase):
    def test_workbook_emits_dot_label_leader_per_point(self) -> None:
        points = _points(4)
        with tempfile.TemporaryDirectory() as temp_dir:
            output = Path(temp_dir) / "sheet.xlsx"

            report = build_sheet(
                [SheetView(image=_image(), points=points)],
                TitleBlock(part_no="64XX2-DR000"),
                output=output,
                template=Path(temp_dir) / "missing.xlsx",
            )
            xml = _drawing_xml(output)

        self.assertEqual(report.pictures, 1)
        self.assertEqual(report.labels, len(points))
        self.assertEqual(report.leaders, len(points))
        self.assertEqual(len(re.findall(r"<pic[ >]", xml)), 1)
        # 세 개가 각각 독립 앵커라야 라벨만 드래그하고 포인트는 제자리에
        # 남길 수 있다. 그룹으로 묶이는 순간 셋이 통째로 움직인다.
        self.assertNotIn("<grpSp", xml)
        self.assertEqual(len(re.findall(r"<sp[ >]", xml)), 2 * len(points))
        self.assertEqual(len(re.findall(r"<cxnSp[ >]", xml)), len(points))

    def test_leader_is_attached_to_label_and_dot(self) -> None:
        # <a:stCxn>/<a:endCxn> 이 있어야 Excel 이 라벨을 옮길 때 지시선을
        # 다시 잇는다. 이 두 태그가 없으면 라벨을 옮겨도 선은 원위치에 남는다.
        points = _points(1)
        with tempfile.TemporaryDirectory() as temp_dir:
            output = Path(temp_dir) / "sheet.xlsx"
            build_sheet(
                [SheetView(image=_image(), points=points)],
                output=output,
                template=Path(temp_dir) / "missing.xlsx",
            )
            xml = _drawing_xml(output)

        leader_xml = re.search(r"<cxnSp\b.*?</cxnSp>", xml, re.DOTALL)
        self.assertIsNotNone(leader_xml)
        body = leader_xml.group(0)
        self.assertRegex(body, r'<a:stCxn id="\d+" idx="\d+"/>')
        self.assertRegex(body, r'<a:endCxn id="\d+" idx="\d+"/>')

    def test_text_box_is_editable(self) -> None:
        # 값 수정이 가능하려면 txBox="1" 이 있어야 한다.
        points = _points(1)
        with tempfile.TemporaryDirectory() as temp_dir:
            output = Path(temp_dir) / "sheet.xlsx"
            build_sheet(
                [SheetView(image=_image(), points=points)],
                output=output,
                template=Path(temp_dir) / "missing.xlsx",
            )
            xml = _drawing_xml(output)
        self.assertIn('txBox="1"', xml)


    def test_values_reach_the_drawing_verbatim(self) -> None:
        points = [SheetPoint("P-01", "-0.7", 0.3, 0.3)]
        with tempfile.TemporaryDirectory() as temp_dir:
            output = Path(temp_dir) / "sheet.xlsx"
            build_sheet(
                [SheetView(image=_image(), points=points)],
                output=output,
                template=Path(temp_dir) / "missing.xlsx",
            )
            xml = _drawing_xml(output)

        self.assertIn("<a:t>-0.7</a:t>", xml)

    def test_title_block_is_written_to_the_form_cells(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            output = Path(temp_dir) / "sheet.xlsx"
            build_sheet(
                [SheetView(image=_image(), points=[])],
                TitleBlock(
                    management_no="ADC-64XX2",
                    part_no="64XX2-DR000",
                    applied_date="2026-08-30",
                ),
                output=output,
                template=Path(temp_dir) / "missing.xlsx",
            )
            sheet = openpyxl.load_workbook(output).active

        self.assertEqual(sheet[config.TITLE_CELLS["management_no"]].value, "ADC-64XX2")
        self.assertEqual(sheet[config.TITLE_CELLS["part_no"]].value, "64XX2-DR000")
        self.assertEqual(sheet[config.TITLE_CELLS["applied_date"]].value, "2026-08-30")

    def test_title_block_style_and_page_break_view_are_explicit(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            output = Path(temp_dir) / "sheet.xlsx"
            build_sheet(
                [SheetView(image=_image(), points=[])],
                TitleBlock(part_no="67XX6", part_name="JD_67XX6-DR000"),
                output=output,
            )
            sheet = openpyxl.load_workbook(output).active

        self.assertEqual(sheet["A1"].value, "보정 적용 내용")
        self.assertEqual(sheet["J1"].value, "관리 NO")
        self.assertEqual(sheet["U1"].value, "PART NAME")
        self.assertEqual(sheet["J3"].value, "공정")
        self.assertEqual(sheet["U3"].value, "PART NO")
        self.assertEqual(sheet["J5"].value, "원소재")
        self.assertEqual(sheet["U5"].value, "적용일자")
        self.assertEqual(sheet["A1"].font.name, "휴먼옛체")
        self.assertEqual(sheet["J1"].font.name, "돋움")
        self.assertEqual(sheet["Y1"].font.name, "맑은 고딕")
        self.assertEqual(sheet["Y3"].font.color.rgb, "FFFF0000")
        self.assertEqual(sheet.sheet_view.view, "pageBreakPreview")
        self.assertEqual(sheet.sheet_view.zoomScale, config.PAGE_BREAK_PREVIEW_ZOOM)
        self.assertFalse(sheet.sheet_view.showGridLines)
        self.assertEqual(str(sheet.print_area), "'00'!$A$1:$AD$40")
        self.assertEqual(sheet.page_setup.orientation, "landscape")
        self.assertEqual(sheet.page_setup.fitToWidth, 1)
        self.assertEqual(sheet.page_setup.fitToHeight, 0)
        self.assertTrue(any(item.id == config.PRINT_PAGE_ROWS for item in sheet.row_breaks.brk))

    def test_custom_title_fonts_cover_heading_labels_and_values(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            output = Path(temp_dir) / "sheet.xlsx"
            build_sheet(
                [SheetView(image=_image(), points=[])],
                TitleBlock(), output=output,
                title_fonts={
                    "heading": "굴림",
                    "management_label": "바탕",
                    "management_no": "Arial",
                },
                title_font_sizes={"heading": 21, "management_label": 9},
            )
            sheet = openpyxl.load_workbook(output).active

        self.assertEqual(sheet["A1"].font.name, "굴림")
        self.assertEqual(sheet["A1"].font.sz, 21)
        self.assertEqual(sheet["J1"].font.name, "바탕")
        self.assertEqual(sheet["J1"].font.sz, 9)
        self.assertEqual(sheet["M1"].font.name, "Arial")

    def test_detail_views_add_their_own_picture_and_caption(self) -> None:
        points = _points(5)
        front = SheetView(image=_image(), points=points)
        detail = crop_view(_image(), points, (0.0, 0.0, 0.5, 1.0), "DETAIL A")
        with tempfile.TemporaryDirectory() as temp_dir:
            output = Path(temp_dir) / "sheet.xlsx"
            report = build_sheet(
                [front, detail], output=output, template=Path(temp_dir) / "missing.xlsx"
            )
            xml = _drawing_xml(output)

        self.assertEqual(report.pictures, 2)
        self.assertEqual(len(re.findall(r"<pic[ >]", xml)), 2)
        self.assertIn("<a:t>DETAIL A</a:t>", xml)

    def test_missing_template_is_reported_but_still_produces_a_file(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            output = Path(temp_dir) / "sheet.xlsx"

            report = build_sheet(
                [SheetView(image=_image(), points=[])],
                output=output,
                template=Path(temp_dir) / "nope.xlsx",
            )

            self.assertTrue(output.is_file())
        self.assertTrue(any("양식" in warning for warning in report.warnings))

    def test_no_views_is_refused(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            with self.assertRaises(ValueError):
                build_sheet([], output=Path(temp_dir) / "sheet.xlsx")

    def test_no_leftover_scratch_file(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            output = Path(temp_dir) / "sheet.xlsx"
            build_sheet(
                [SheetView(image=_image(), points=_points(2))],
                output=output,
                template=Path(temp_dir) / "missing.xlsx",
            )

            leftovers = [p.name for p in Path(temp_dir).iterdir() if p != output]

        self.assertEqual(leftovers, [])


if __name__ == "__main__":
    unittest.main()

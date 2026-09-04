"""Assemble the correction sheet workbook.

openpyxl handles the cells and the pictures -- including the media parts and
relationships -- and the value labels are appended to the drawing part
afterwards, because openpyxl cannot create text boxes or connectors.
"""

from __future__ import annotations

import io
import re
import zipfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Sequence

import cv2
import numpy as np
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import AbsoluteAnchor
from openpyxl.drawing.xdr import XDRPoint2D, XDRPositiveSize2D
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.pagebreak import Break
from openpyxl.utils import range_boundaries

from . import config, drawing
from .layout import SheetView, default_layout, place_labels


@dataclass
class TitleBlock:
    """The values that go into the sheet's header cells."""

    heading: str = "보정 적용 내용"
    management_label: str = "관리 NO"
    management_no: str = ""
    part_name_label: str = "PART NAME"
    part_name: str = ""
    process_label: str = "공정"
    process: str = ""
    part_no_label: str = "PART NO"
    part_no: str = ""
    material_label: str = "원소재"
    material: str = ""
    applied_date_label: str = "적용일자"
    applied_date: str = ""

    def as_cells(self) -> dict[str, str]:
        return {
            cell: getattr(self, field_name)
            for field_name, cell in config.TITLE_CELLS.items()
            if getattr(self, field_name)
        }


@dataclass
class BuildReport:
    """What actually landed on the sheet."""

    path: Path
    pictures: int = 0
    labels: int = 0
    leaders: int = 0
    warnings: list[str] = field(default_factory=list)


# Excel-observed connection-site index mapping. For text boxes the
# left/right indices come out opposite of ECMA-376 -- attaching the leader
# to idx=3 lands it on the right edge of the box (crossing through the
# label) instead of the left. Empirically Excel treats a text box as
#   0 = top, 1 = left, 2 = bottom, 3 = right.
# Ellipse (the anchor dot) follows the spec:
#   0 = top, 1 = right, 2 = bottom, 3 = left.
_LABEL_IDX = {"top": 0, "right": 3, "bottom": 2, "left": 1}
_DOT_IDX = {"top": 0, "right": 1, "bottom": 2, "left": 3}

# Given the edge the label sits on, which side of the label the leader
# exits and which side of the dot it enters.
_LEADER_SIDES = {
    "top":    ("bottom", "top"),     # label above point
    "bottom": ("top",    "bottom"),  # label below point
    "left":   ("right",  "left"),    # label to the left of point
    "right":  ("left",   "right"),   # label to the right of point
}


def _connection_setup(edge: str) -> tuple[str, str, int, int]:
    """Return (label_side, dot_side, label_idx, dot_idx) for an edge."""
    label_side, dot_side = _LEADER_SIDES.get(edge, ("bottom", "top"))
    return label_side, dot_side, _LABEL_IDX[label_side], _DOT_IDX[dot_side]


def _side_center(
    x: float, y: float, width: float, height: float, side: str,
) -> tuple[float, float]:
    """Coordinate at the middle of one named edge of a rectangle."""
    if side == "top":
        return x + width / 2, y
    if side == "right":
        return x + width, y + height / 2
    if side == "bottom":
        return x + width / 2, y + height
    if side == "left":
        return x, y + height / 2
    raise ValueError(f"모르는 side 이름: {side!r}")


def _encode_png(image: np.ndarray) -> io.BytesIO:
    ok, encoded = cv2.imencode(".png", image)
    if not ok:
        raise ValueError("뷰 이미지를 PNG로 변환하지 못했습니다.")
    return io.BytesIO(encoded.tobytes())


def _merge_title_block(sheet) -> None:
    """Rebuild every title merge so a damaged/missing template is harmless."""
    min_col, min_row, max_col, max_row = range_boundaries("A1:AD6")
    overlapping = [
        merged
        for merged in list(sheet.merged_cells.ranges)
        if not (
            merged.max_col < min_col
            or merged.min_col > max_col
            or merged.max_row < min_row
            or merged.min_row > max_row
        )
    ]
    for merged in overlapping:
        sheet.unmerge_cells(str(merged))
    for merged_range in config.TITLE_BLOCK_MERGES:
        sheet.merge_cells(merged_range)


def _title_font(field_name: str, title_fonts, title_font_sizes) -> Font:
    family = (title_fonts or {}).get(field_name) or config.TITLE_DEFAULT_FONTS[field_name]
    size = (title_font_sizes or {}).get(field_name)
    try:
        size = float(size) if size is not None else config.TITLE_DEFAULT_SIZES[field_name]
    except (TypeError, ValueError):
        size = config.TITLE_DEFAULT_SIZES[field_name]
    return Font(
        name=str(family), size=size,
        bold=field_name in {"heading", "part_no"},
        color="FFFF0000" if field_name == "part_no" else "FF000000",
        charset=129,
    )


def _style_title_block(sheet, title: TitleBlock, title_fonts=None,
                       title_font_sizes=None) -> None:
    """Apply the company title box without relying on template side effects."""
    _merge_title_block(sheet)
    cells = {
        "heading": config.SHEET_HEADING_ANCHOR,
        **config.TITLE_LABEL_CELLS,
        **config.TITLE_CELLS,
    }
    for field_name, cell_name in cells.items():
        cell = sheet[cell_name]
        cell.value = getattr(title, field_name)
        cell.font = _title_font(field_name, title_fonts, title_font_sizes)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    thin = Side(style="thin", color="FF000000")
    white = PatternFill(fill_type="solid", fgColor="FFFFFFFF")
    for row in sheet.iter_rows(min_row=1, max_row=6, min_col=1, max_col=30):
        for cell in row:
            cell.fill = white
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(horizontal="center", vertical="center")


def _apply_print_layout(sheet, max_row: int = config.PRINT_PAGE_ROWS) -> None:
    """Open in Page Break Preview and print one landscape page per 40 rows."""
    sheet.sheet_view.view = "pageBreakPreview"
    sheet.sheet_view.zoomScale = config.PAGE_BREAK_PREVIEW_ZOOM
    sheet.sheet_view.zoomScaleNormal = 100
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = None
    sheet.print_area = f"A1:AD{max_row}"
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.sheet_properties.pageSetUpPr.autoPageBreaks = False
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.page_setup.scale = None
    sheet.page_margins.left = 0
    sheet.page_margins.right = 0
    sheet.page_margins.top = 0
    sheet.page_margins.bottom = 0
    sheet.print_options.horizontalCentered = True
    sheet.print_options.verticalCentered = False
    sheet.row_breaks = type(sheet.row_breaks)()
    sheet.row_breaks.append(Break(id=config.PRINT_PAGE_ROWS))


def _open_template(template: Path | None) -> openpyxl.Workbook:
    """Use the company form when it is available, otherwise a blank book.

    openpyxl drops the form's own drawing and printer settings on save, so the
    blank fallback keeps the tool usable rather than matching the form exactly.
    """
    candidate = Path(template) if template else config.DEFAULT_TEMPLATE
    if candidate.is_file():
        return openpyxl.load_workbook(candidate)
    return openpyxl.Workbook()


def build_sheet(
    views: Sequence[SheetView],
    title: TitleBlock | None = None,
    *,
    output: Path,
    template: Path | None = None,
    title_fonts: dict[str, str] | None = None,
    title_font_sizes: dict[str, float | int | None] | None = None,
    point_font_family: str | None = None,
) -> BuildReport:
    """Write the correction sheet and report what it contains."""
    if not views:
        raise ValueError("시트에 올릴 뷰가 없습니다.")

    output = Path(output)
    output.parent.mkdir(parents=True, exist_ok=True)
    warnings: list[str] = []

    resolved = Path(template) if template else config.DEFAULT_TEMPLATE
    if not resolved.is_file():
        warnings.append(
            f"양식 파일이 없어 빈 통합문서로 만들었습니다: {resolved}"
        )

    book = _open_template(template)
    sheet = book.active
    title_values = title or TitleBlock()
    _style_title_block(sheet, title_values, title_fonts, title_font_sizes)
    _apply_print_layout(sheet)

    default_layout(list(views))

    # The UI annotation layer spans the whole preview canvas, not just the
    # part picture, so annotation ratios are % of that canvas. Mirror it
    # here by allocating the sheet's drawing area (from the row below the
    # title block to the bottom guide) as the annotation canvas — that way
    # a rectangle drawn in the margin around the picture in the UI still
    # lands in the margin around the picture in the sheet.
    annotation_canvas = (
        0.0, float(config.DRAWING_TOP),
        float(config.SHEET_WIDTH),
        float(config.DRAWING_BOTTOM - config.DRAWING_TOP),
    )

    anchors: list[str] = []
    shape_id = 1000
    labels = leaders = 0
    for view_index, view in enumerate(views):
        box_x, box_y, box_width, box_height = view.box
        picture = XLImage(_encode_png(view.image))
        picture.anchor = AbsoluteAnchor(
            pos=XDRPoint2D(drawing.emu(box_x), drawing.emu(box_y)),
            ext=XDRPositiveSize2D(drawing.emu(box_width), drawing.emu(box_height)),
        )
        sheet.add_image(picture)

        if view.title:
            shape_id += 1
            anchors.append(
                drawing.caption(
                    shape_id,
                    view.title,
                    box_x,
                    box_y - config.DETAIL_TITLE_HEIGHT - 2,
                    box_width,
                )
            )

        for placed in place_labels(view):
            # Three separate anchors per callout: dot, label, leader. The
            # leader carries <a:stCxn>/<a:endCxn> so Excel treats it as an
            # *attached* connector -- drag the label and the line stretches
            # to stay connected to the fixed measurement point. Grouping was
            # tried first but locked all three into a rigid unit, which
            # moved the dot away from the real measurement location.
            label_id = shape_id + 1
            dot_id = shape_id + 2
            leader_id = shape_id + 3
            shape_id += 3

            anchors.append(
                drawing.dot(
                    dot_id, placed.point_x, placed.point_y,
                    config.POINT_DOT_RADIUS,
                )
            )
            anchors.append(
                drawing.text_box(
                    label_id, placed.text, placed.label_x, placed.label_y,
                    font_family=point_font_family,
                )
            )

            # Attach the connector to whichever side of the label faces the
            # point at placement time. Excel keeps this attachment through
            # future moves, so the leader always emits from that same side
            # of the label. The initial x1/y1 must match that side too --
            # centring the initial line on the label draws it visibly
            # crossing the box until Excel recomputes on first save.
            label_side_name, _dot_side_name, label_idx, dot_idx = (
                _connection_setup(placed.edge)
            )
            label_start_x, label_start_y = _side_center(
                placed.label_x, placed.label_y,
                config.LABEL_WIDTH, config.LABEL_HEIGHT,
                label_side_name,
            )
            anchors.append(
                drawing.leader(
                    leader_id,
                    label_start_x, label_start_y,
                    placed.point_x, placed.point_y,
                    start_shape_id=label_id, start_idx=label_idx,
                    end_shape_id=dot_id, end_idx=dot_idx,
                )
            )
            leaders += 1
            labels += 1

        # Free-form annotations from the UI. Their ratios are of the whole
        # canvas, so they use ``annotation_canvas`` (the sheet's drawing
        # area) rather than the picture's fit_box, and only the front view
        # renders them -- detail views crop the picture but not the canvas.
        if view_index == 0:
            for annotation in view.annotations:
                shape_id += 1
                anchor = _annotation_anchor(
                    shape_id, annotation, annotation_canvas,
                )
                if anchor is not None:
                    anchors.append(anchor)

    book.save(output)
    _append_anchors(output, anchors)

    return BuildReport(
        path=output,
        pictures=len(views),
        labels=labels,
        leaders=leaders,
        warnings=warnings,
    )


def _annotation_anchor(
    shape_id: int,
    annotation,
    view_box: tuple[float, float, float, float],
) -> str | None:
    """Turn one UI annotation into the drawing anchor that renders it."""
    box_x, box_y, box_width, box_height = view_box
    x = box_x + annotation.x_ratio * box_width
    y = box_y + annotation.y_ratio * box_height
    width = annotation.w_ratio * box_width
    height = annotation.h_ratio * box_height

    if annotation.kind == "arrow":
        # Arrow encodes direction in the sign of w/h; keep both endpoints
        # so a right-to-left arrow still points at the right target.
        return drawing.annotation_arrow(
            shape_id, x, y, x + width, y + height, annotation.color,
        )

    # For box-shaped annotations flip negative extents so the outer
    # <pos>/<ext> stay positive as Excel requires.
    if width < 0:
        x += width
        width = -width
    if height < 0:
        y += height
        height = -height
    # Sub-pixel or zero-size shapes are invisible; treat them like the UI
    # does and clamp to at least a couple of pixels.
    width = max(width, 2.0)
    height = max(height, 2.0)

    if annotation.kind == "rect":
        return drawing.annotation_rect(
            shape_id, x, y, width, height, annotation.color,
        )
    if annotation.kind == "ellipse":
        return drawing.annotation_ellipse(
            shape_id, x, y, width, height, annotation.color,
        )
    if annotation.kind == "text":
        text = annotation.text or ""
        if not text.strip():
            return None
        # The UI stores font size as CSS pixels relative to the layer, which
        # scales 1:1 with the preview image. The Excel picture uses the same
        # pixel frame as its placed box, so use those pixels directly and
        # convert to points (Excel's text unit) at 1px == 0.75pt.
        size_px = annotation.font_size_px or 10.0
        return drawing.annotation_text(
            shape_id, text, x, y, width, height,
            annotation.color, size_px * 0.75,
        )
    return None


# 13.5pt row height -> 18px. Used to translate the "N rows down" offset the
# stacked-append feature applies to the new block's images and shapes.
_ROW_HEIGHT_PX = 18


def stack_workbooks(previous_bytes: bytes, new_bytes: bytes) -> bytes:
    """Append the new single-block workbook onto the previous one.

    The previous workbook is kept intact (its text boxes, connectors, and
    dots that openpyxl cannot round-trip are preserved because we operate on
    the raw zip). The new workbook's cells, drawings, and media are shifted
    down by the previous block's row count and appended, so the two blocks
    print as consecutive pages of one sheet.

    Assumes the workbook layout ``build_sheet`` produces: one worksheet
    (``xl/worksheets/sheet1.xml``), one drawing part
    (``xl/drawings/drawing1.xml``), pictures under ``xl/media/``. Anything
    outside that shape is copied straight from the previous file, so a hand
    edited template will not be re-flowed.
    """
    with zipfile.ZipFile(io.BytesIO(previous_bytes)) as prev_zip, \
            zipfile.ZipFile(io.BytesIO(new_bytes)) as new_zip:
        prev_names = set(prev_zip.namelist())
        prev_sheet_name = _first_matching(prev_names, _SHEET_XML_RE)
        prev_drawing_name = _first_matching(prev_names, drawing.DRAWING_PART)
        if prev_sheet_name is None or prev_drawing_name is None:
            raise ValueError(
                "이전 시트에서 워크시트 또는 drawing 파트를 찾지 못했습니다."
            )
        prev_sheet_xml = prev_zip.read(prev_sheet_name).decode("utf-8")
        prev_drawing_xml = prev_zip.read(prev_drawing_name).decode("utf-8")
        prev_drawing_rels_name = _rels_for(prev_drawing_name)
        prev_drawing_rels = (
            prev_zip.read(prev_drawing_rels_name).decode("utf-8")
            if prev_drawing_rels_name in prev_names
            else _EMPTY_RELS
        )
        prev_row_offset = _sheet_max_row(prev_sheet_xml)
        prev_max_rid = _max_relationship_id(prev_drawing_rels)
        prev_media_ids = _existing_media_ids(prev_names)

        new_sheet_name = _first_matching(set(new_zip.namelist()), _SHEET_XML_RE)
        new_drawing_name = _first_matching(
            set(new_zip.namelist()), drawing.DRAWING_PART
        )
        if new_sheet_name is None or new_drawing_name is None:
            raise ValueError(
                "새 시트에서 워크시트 또는 drawing 파트를 찾지 못했습니다."
            )
        new_sheet_xml = new_zip.read(new_sheet_name).decode("utf-8")
        new_drawing_xml = new_zip.read(new_drawing_name).decode("utf-8")
        new_drawing_rels_name = _rels_for(new_drawing_name)
        new_drawing_rels = (
            new_zip.read(new_drawing_rels_name).decode("utf-8")
            if new_drawing_rels_name in set(new_zip.namelist())
            else _EMPTY_RELS
        )

        # Renumber the new block's images so they do not clash with the
        # previous block's rIds and media file names.
        rid_remap, media_remap = _plan_media_remap(
            new_drawing_rels, prev_max_rid, prev_media_ids
        )
        remapped_new_drawing_rels = _remap_rels(new_drawing_rels, rid_remap, media_remap)
        remapped_new_drawing_xml = _remap_embeds(new_drawing_xml, rid_remap)

        # Shape ids (cNvPr id) must also be unique across the merged drawing,
        # otherwise Excel treats the file as corrupt on open and drops the
        # duplicate anchors during repair. The attached connectors reference
        # the label/dot ids via <a:stCxn>/<a:endCxn>, so those get remapped
        # in lockstep to keep the leaders attached to the right shapes.
        prev_max_shape_id = _max_shape_id(prev_drawing_xml)
        shape_id_remap = _plan_shape_id_remap(
            remapped_new_drawing_xml, prev_max_shape_id
        )
        remapped_new_drawing_xml = _remap_shape_ids(
            remapped_new_drawing_xml, shape_id_remap
        )

        pixel_offset = prev_row_offset * _ROW_HEIGHT_PX
        shifted_new_sheet = _shift_sheet_xml(new_sheet_xml, prev_row_offset)
        shifted_new_drawing = _shift_drawing_anchors(
            remapped_new_drawing_xml, pixel_offset
        )

        merged_sheet_xml = _merge_sheet_xml(
            prev_sheet_xml, shifted_new_sheet, prev_row_offset
        )
        merged_drawing_xml = _merge_drawing_xml(prev_drawing_xml, shifted_new_drawing)
        merged_drawing_rels = _merge_rels(prev_drawing_rels, remapped_new_drawing_rels)

        # Expand <dimension> and the print-area defined name to cover the
        # appended rows. Without this the template's fixed A1:AD40 print
        # area leaves every stacked block outside it -- in pageBreakPreview
        # that whole region renders grey, so the labels (drawn on a white
        # background) become nearly invisible and the block reads as
        # "missing annotations".
        merged_max_row = _sheet_max_row(merged_sheet_xml)
        merged_sheet_xml = _expand_dimension(merged_sheet_xml, merged_max_row)

        workbook_xml_name = "xl/workbook.xml"
        merged_workbook_xml: str | None = None
        if workbook_xml_name in prev_names:
            merged_workbook_xml = _expand_print_area(
                prev_zip.read(workbook_xml_name).decode("utf-8"),
                merged_max_row,
            )

        buffer = io.BytesIO()
        with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as out:
            written: set[str] = set()
            for info in prev_zip.infolist():
                if info.filename == prev_sheet_name:
                    out.writestr(info, merged_sheet_xml.encode("utf-8"))
                elif info.filename == prev_drawing_name:
                    out.writestr(info, merged_drawing_xml.encode("utf-8"))
                elif prev_drawing_rels_name and info.filename == prev_drawing_rels_name:
                    out.writestr(info, merged_drawing_rels.encode("utf-8"))
                elif (
                    merged_workbook_xml is not None
                    and info.filename == workbook_xml_name
                ):
                    out.writestr(info, merged_workbook_xml.encode("utf-8"))
                else:
                    out.writestr(info, prev_zip.read(info.filename))
                written.add(info.filename)
            # Copy new media files with their remapped names.
            for old_name, new_name in media_remap.items():
                if new_name in written:
                    continue
                out.writestr(new_name, new_zip.read(old_name))
                written.add(new_name)
        return buffer.getvalue()


_SHEET_XML_RE = re.compile(r"^xl/worksheets/sheet\d+\.xml$")
_EMPTY_RELS = (
    '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
    '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"/>'
)


def _first_matching(names: set[str], pattern: re.Pattern[str]) -> str | None:
    for name in names:
        if pattern.match(name) or pattern.fullmatch(name):
            return name
    return None


def _rels_for(part_name: str) -> str:
    directory, _, filename = part_name.rpartition("/")
    return f"{directory}/_rels/{filename}.rels"


def _sheet_max_row(sheet_xml: str) -> int:
    rows = re.findall(r'<row\b[^>]*\br="(\d+)"', sheet_xml)
    return max((int(r) for r in rows), default=0)


def _expand_dimension(sheet_xml: str, max_row: int) -> str:
    """Widen ``<dimension ref>`` so its bottom row reaches ``max_row``."""
    if max_row <= 0:
        return sheet_xml

    def _replace(match: re.Match[str]) -> str:
        prefix, ref, suffix = match.group(1), match.group(2), match.group(3)
        parts = ref.split(":")
        first = parts[0]
        last = parts[1] if len(parts) > 1 else parts[0]
        last_col_match = re.match(r"^([A-Z]+)(\d+)$", last)
        if last_col_match is None:
            return match.group(0)
        column, row = last_col_match.group(1), int(last_col_match.group(2))
        if row >= max_row:
            return match.group(0)
        return f'{prefix}"{first}:{column}{max_row}"{suffix}'

    return re.sub(
        r'(<dimension\b[^>]*\bref=)"([A-Z]+\d+(?::[A-Z]+\d+)?)"([^/>]*/?>)',
        _replace, sheet_xml, count=1,
    )


def _expand_print_area(workbook_xml: str, max_row: int) -> str:
    """Grow every ``_xlnm.Print_Area`` defined name to include ``max_row``.

    The template ships with ``'00'!$A$1:$AD$40``; without expansion the
    stacked pages sit outside the print area and pageBreakPreview greys
    them out so the labels blend into the background.
    """
    if max_row <= 0:
        return workbook_xml

    def _replace(match: re.Match[str]) -> str:
        opening, body, closing = match.group(1), match.group(2), match.group(3)
        expanded = re.sub(
            r"(\$?[A-Z]+\$?\d+:\$?[A-Z]+\$?)(\d+)",
            lambda m: (
                f"{m.group(1)}{max_row}"
                if int(m.group(2)) < max_row else m.group(0)
            ),
            body,
        )
        return f"{opening}{expanded}{closing}"

    return re.sub(
        r'(<definedName\b[^>]*\bname="_xlnm\.Print_Area"[^>]*>)(.*?)(</definedName>)',
        _replace, workbook_xml, flags=re.DOTALL,
    )


def _shift_cell_ref(ref: str, row_offset: int) -> str:
    match = re.match(r"^([A-Z]+)(\d+)$", ref)
    if not match:
        return ref
    return f"{match.group(1)}{int(match.group(2)) + row_offset}"


def _shift_range_ref(ref: str, row_offset: int) -> str:
    return ":".join(_shift_cell_ref(part, row_offset) for part in ref.split(":"))


def _shift_sheet_xml(sheet_xml: str, row_offset: int) -> str:
    """Shift every row/cell reference in the sheet XML down by row_offset."""
    if row_offset <= 0:
        return sheet_xml

    def _row(match: re.Match[str]) -> str:
        return f'{match.group(1)}"{int(match.group(2)) + row_offset}"'

    xml = re.sub(r'(<row\b[^>]*\br=)"(\d+)"', _row, sheet_xml)

    def _cell(match: re.Match[str]) -> str:
        return f'{match.group(1)}"{_shift_cell_ref(match.group(2), row_offset)}"'

    xml = re.sub(r'(<c\b[^>]*\br=)"([A-Z]+\d+)"', _cell, xml)

    def _range(match: re.Match[str]) -> str:
        return f'{match.group(1)}"{_shift_range_ref(match.group(2), row_offset)}"'

    # Ranges show up on <mergeCell ref>, <mergeCells>, <printArea>, etc.
    xml = re.sub(
        r'(\bref=)"([A-Z]+\d+(?::[A-Z]+\d+)?)"', _range, xml
    )
    return xml


def _shift_drawing_anchors(drawing_xml: str, pixel_offset_px: int) -> str:
    """Shift y coordinates of anchor positions and shape offsets."""
    if pixel_offset_px <= 0:
        return drawing_xml
    emu = pixel_offset_px * config.EMU_PER_PIXEL

    def _pos_y(match: re.Match[str]) -> str:
        return f'{match.group(1)}"{int(match.group(2)) + emu}"'

    xml = re.sub(r'(<pos\b[^>]*\by=)"(\d+)"', _pos_y, drawing_xml)

    def _off_y(match: re.Match[str]) -> str:
        return f'{match.group(1)}"{int(match.group(2)) + emu}"'

    xml = re.sub(r'(<a:off\b[^>]*\by=)"(\d+)"', _off_y, xml)
    return xml


def _max_relationship_id(rels_xml: str) -> int:
    ids = re.findall(r'Id="rId(\d+)"', rels_xml)
    return max((int(i) for i in ids), default=0)


def _existing_media_ids(names: set[str]) -> set[int]:
    ids: set[int] = set()
    for name in names:
        match = re.match(r"xl/media/image(\d+)\.\w+$", name)
        if match:
            ids.add(int(match.group(1)))
    return ids


def _plan_media_remap(
    new_rels_xml: str, prev_max_rid: int, prev_media_ids: set[int]
) -> tuple[dict[str, str], dict[str, str]]:
    """Pick new rIds and image file names for each of the new block's images.

    Extract Id and Target with separate regexes because openpyxl and Excel
    write the attributes in different orders (openpyxl: Type/Target/Id,
    Excel: Id/Type/Target), so a single order-sensitive regex misses one of
    the two forms and the merge silently reuses the previous block's image.
    """
    rid_remap: dict[str, str] = {}
    media_remap: dict[str, str] = {}
    next_rid = prev_max_rid
    next_media = max(prev_media_ids, default=0)
    # ``[^>]*?`` non-greedy up to ``/>``. Earlier ``[^/]*`` broke on URLs
    # like Type="http://schemas..." which contain slashes.
    for elem_match in re.finditer(r"<Relationship\b[^>]*?/>", new_rels_xml):
        text = elem_match.group(0)
        id_match = re.search(r'Id="(rId\d+)"', text)
        target_match = re.search(r'Target="([^"]+)"', text)
        if id_match is None or target_match is None:
            continue
        media_match = re.search(r"media/image(\d+)\.(\w+)$", target_match.group(1))
        if media_match is None:
            continue
        next_rid += 1
        rid_remap[id_match.group(1)] = f"rId{next_rid}"
        next_media += 1
        old_file = f"xl/media/image{media_match.group(1)}.{media_match.group(2)}"
        media_remap[old_file] = (
            f"xl/media/image{next_media}.{media_match.group(2)}"
        )
    return rid_remap, media_remap


def _remap_rels(
    rels_xml: str, rid_remap: dict[str, str], media_remap: dict[str, str]
) -> str:
    xml = rels_xml
    for old, new in rid_remap.items():
        xml = xml.replace(f'Id="{old}"', f'Id="{new}"')
    for old_full, new_full in media_remap.items():
        media_tail = old_full.split("xl/", 1)[1]
        new_tail = new_full.split("xl/", 1)[1]
        # openpyxl writes Target as "/xl/media/imageN.png" (absolute), but
        # Excel-authored files use "../media/imageN.png". Cover both so the
        # relationship still points at the renamed file.
        for old_form, new_form in (
            (f"/xl/{media_tail}", f"/xl/{new_tail}"),
            (f"../{media_tail}", f"../{new_tail}"),
            (f"/{media_tail}", f"/{new_tail}"),
        ):
            xml = xml.replace(f'Target="{old_form}"', f'Target="{new_form}"')
    return xml


def _remap_embeds(drawing_xml: str, rid_remap: dict[str, str]) -> str:
    xml = drawing_xml
    for old, new in rid_remap.items():
        xml = xml.replace(f'r:embed="{old}"', f'r:embed="{new}"')
    return xml


def _max_shape_id(drawing_xml: str) -> int:
    ids = re.findall(r'<cNvPr\b[^/>]*\bid="(\d+)"', drawing_xml)
    return max((int(i) for i in ids), default=0)


def _plan_shape_id_remap(drawing_xml: str, id_offset: int) -> dict[str, str]:
    """Give every cNvPr id in the new drawing a value above ``id_offset``."""
    remap: dict[str, str] = {}
    next_id = id_offset
    for match in re.finditer(r'<cNvPr\b[^/>]*\bid="(\d+)"', drawing_xml):
        old = match.group(1)
        if old in remap:
            continue
        next_id += 1
        remap[old] = str(next_id)
    return remap


def _remap_shape_ids(drawing_xml: str, id_remap: dict[str, str]) -> str:
    if not id_remap:
        return drawing_xml

    def _cnvpr(match: re.Match[str]) -> str:
        prefix, old, suffix = match.group(1), match.group(2), match.group(3)
        return f'{prefix}{id_remap.get(old, old)}{suffix}'

    xml = re.sub(
        r'(<cNvPr\b[^/>]*\bid=")(\d+)(")', _cnvpr, drawing_xml
    )

    def _cxn(match: re.Match[str]) -> str:
        prefix, old, suffix = match.group(1), match.group(2), match.group(3)
        return f'{prefix}{id_remap.get(old, old)}{suffix}'

    xml = re.sub(
        r'(<a:(?:st|end)Cxn\b[^/>]*\bid=")(\d+)(")', _cxn, xml
    )
    return xml


def _merge_sheet_xml(prev_xml: str, shifted_new_xml: str, prev_row_offset: int) -> str:
    """Append the shifted new block's rows and mergeCells into the previous sheet."""
    new_rows = _extract_between(
        shifted_new_xml, "<sheetData>", "</sheetData>"
    ) or _extract_self_closed(shifted_new_xml, "sheetData")
    new_merges = _extract_between(
        shifted_new_xml, "<mergeCells", "</mergeCells>",
        include_start=True,
    )

    xml = prev_xml
    if new_rows:
        # Insert new rows just before </sheetData>
        xml = xml.replace("</sheetData>", new_rows + "</sheetData>", 1)
    if new_merges:
        merge_items_match = re.search(r"<mergeCells[^>]*>(.*?)</mergeCells>", new_merges, re.DOTALL)
        if merge_items_match:
            items = merge_items_match.group(1)
            if "<mergeCells" in xml:
                # Append inside existing mergeCells and bump count
                xml = re.sub(
                    r"(<mergeCells[^>]*count=)\"(\d+)\"([^>]*>)(.*?)(</mergeCells>)",
                    lambda m: (
                        f"{m.group(1)}\"{int(m.group(2)) + items.count('<mergeCell ')}\""
                        f"{m.group(3)}{m.group(4)}{items}{m.group(5)}"
                    ),
                    xml,
                    count=1,
                    flags=re.DOTALL,
                )
            else:
                # Add fresh mergeCells right after </sheetData>
                xml = xml.replace(
                    "</sheetData>",
                    f"</sheetData><mergeCells count=\"{items.count('<mergeCell ')}\">{items}</mergeCells>",
                    1,
                )
    # Add a page break at the previous block's last row so printing splits cleanly.
    xml = _ensure_row_break(xml, prev_row_offset)
    return xml


def _extract_between(
    xml: str, start: str, end: str, *, include_start: bool = False
) -> str:
    start_index = xml.find(start)
    if start_index < 0:
        return ""
    end_index = xml.find(end, start_index)
    if end_index < 0:
        return ""
    if include_start:
        return xml[start_index:end_index + len(end)]
    return xml[start_index + len(start):end_index]


def _extract_self_closed(xml: str, tag: str) -> str:
    match = re.search(rf"<{tag}\b[^/]*/>", xml)
    return "" if match is None else ""


def _ensure_row_break(sheet_xml: str, break_row: int) -> str:
    if break_row <= 0:
        return sheet_xml
    row_break = (
        f'<brk id="{break_row}" max="16383" man="1"/>'
    )
    if "<rowBreaks" in sheet_xml:
        return re.sub(
            r"(<rowBreaks[^>]*count=)\"(\d+)\"([^>]*manualBreakCount=)\"(\d+)\"([^>]*>)",
            lambda m: (
                f"{m.group(1)}\"{int(m.group(2)) + 1}\""
                f"{m.group(3)}\"{int(m.group(4)) + 1}\"{m.group(5)}{row_break}"
            ),
            sheet_xml,
            count=1,
        )
    row_breaks = (
        f'<rowBreaks count="1" manualBreakCount="1">{row_break}</rowBreaks>'
    )
    # Per CT_Worksheet: rowBreaks belongs AFTER pageSetup/headerFooter and
    # BEFORE drawing. Inserting it earlier (which openpyxl-style parsers
    # tolerate but Excel does not) makes the file unopenable. Anchor to
    # <drawing so we land after every optional page-setup element that may
    # or may not be present.
    for anchor in ("<drawing ", "<drawing/>", "<legacyDrawing", "</worksheet>"):
        if anchor in sheet_xml:
            return sheet_xml.replace(anchor, row_breaks + anchor, 1)
    return sheet_xml


def _merge_drawing_xml(prev_xml: str, shifted_new_xml: str) -> str:
    new_anchors = _extract_between(shifted_new_xml, "", "</wsDr>")
    # Get everything inside <wsDr>...</wsDr> from the new drawing
    start = shifted_new_xml.find(">", shifted_new_xml.find("<wsDr")) + 1
    end = shifted_new_xml.find("</wsDr>")
    if start <= 0 or end <= 0 or start >= end:
        return prev_xml
    new_body = shifted_new_xml[start:end]
    if "</wsDr>" not in prev_xml:
        return prev_xml
    return prev_xml.replace("</wsDr>", new_body + "</wsDr>", 1)


def _merge_rels(prev_rels: str, new_rels: str) -> str:
    new_relationships = re.findall(r"<Relationship\b[^>]*?/>", new_rels)
    if not new_relationships:
        return prev_rels
    joined = "".join(new_relationships)
    if "</Relationships>" not in prev_rels:
        return prev_rels
    return prev_rels.replace("</Relationships>", joined + "</Relationships>", 1)


def _append_anchors(path: Path, anchors: list[str]) -> None:
    """Rewrite the workbook zip with the extra shapes in the drawing part.

    The scratch file sits next to the output so the final move stays on one
    filesystem and no temporary handle is left open on Windows.
    """
    if not anchors:
        return
    temporary = path.with_name(path.name + ".building")
    temporary.unlink(missing_ok=True)
    try:
        injected = False
        with zipfile.ZipFile(path) as source, zipfile.ZipFile(
            temporary, "w", zipfile.ZIP_DEFLATED
        ) as target:
            for info in source.infolist():
                payload = source.read(info.filename)
                if drawing.DRAWING_PART.match(info.filename):
                    payload = drawing.inject(
                        payload.decode("utf-8"), anchors
                    ).encode("utf-8")
                    injected = True
                target.writestr(info, payload)
        if not injected:
            raise ValueError("이미지가 없어 라벨을 넣을 drawing 파트를 찾지 못했습니다.")
        temporary.replace(path)
    finally:
        temporary.unlink(missing_ok=True)

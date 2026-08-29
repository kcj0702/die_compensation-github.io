"""보정시트를 현업 엑셀 양식으로 낸다.

[양식을 뜯어보고 알게 된 것]
현업이 준 `보정시트_양식.xlsx` 와 실제 작성 사례 두 건을 열어 봤다.

    보정시트_양식.xlsx          40행 x 30열, 인쇄영역 A1:AD40, A4 가로
    CD8 71XX2_22 보정내용.xlsx  이미지 10개, 값 있는 칸 48
    JM 67312-DZ000_보정적용.xlsx 이미지 7개,  값 있는 칸 40

**표가 아니라 그림이다.** 머리말 여섯 칸(관리 NO / PART NAME / 공정 /
PART NO / 원소재 / 적용일자)만 글자고, 나머지는 보정치를 그려 넣은
스캔 그림을 통째로 붙인다. 그리고 그 40행 묶음이 공정마다 반복된다
(1행, 41행, 81행 …).

    J1:L2 관리 NO   M1:T2 (값)   U1:X2 PART NAME  Y1:AD2 (값)
    J3:L4 공   정   M3:T4 (값)   U3:X4 PART NO    Y3:AD4 (값)
    J5:L6 원소재    M5:T6 (값)   U5:X6 적용일자   Y5:AD6 (값)

[그래서 이렇게 만든다]
1쪽 — 양식 그대로. 머리말을 채우고 보정치를 그린 그림을 붙인다.
2쪽 — 포인트 표. 원래 양식에는 없지만 "포인트 별 편차 엑셀 자동 작성"
      이 향후 계획 항목이라 따로 시트를 붙인다. 사람이 그림을 보고,
      기계가 표를 읽는다.

[그림은 여기서 그린다]
화면(DOM)을 이미지로 굽는 방법도 있지만 html2canvas 같은 걸 새로
받아야 한다. 사내망에서 도는 게 이 프로젝트의 전제라 서버에서
OpenCV 로 직접 그린다.
"""
from __future__ import annotations

import io
from dataclasses import dataclass
from datetime import date
from pathlib import Path

import cv2
import numpy as np

TEMPLATE = Path(__file__).resolve().parent / "templates" / "보정시트_양식.xlsx"

# 양식의 머리말 자리 — 왼쪽 라벨은 이미 적혀 있고 값 칸만 채운다.
HEADER_CELLS = {
    "control_no": "M1",
    "part_name": "Y1",
    "process": "M3",
    "part_no": "Y3",
    "material": "M5",
    "applied_at": "Y5",
}
# 그림이 들어갈 자리. 머리말이 6행까지라 8행부터 쓴다.
IMAGE_ANCHOR = "A8"
IMAGE_ROWS = 32           # 8행부터 40행까지
ROW_POINTS = 14.0         # 양식의 행 높이(pt)

PLUS = (63, 72, 224)      # BGR — 살이 많다(깎는다)
MINUS = (224, 127, 47)    # BGR — 살이 부족하다(붙인다)


@dataclass
class SheetPoint:
    """보정시트에 찍힐 포인트 하나."""

    point_id: str
    x_px: int
    y_px: int
    deviation: float      # 스캔에서 읽은 편차
    correction: float     # 최종 보정량 (작업자 수정 반영)


def draw_sheet_image(base_bgr: np.ndarray, points: list) -> np.ndarray:
    """스캔 그림 위에 보정치를 콜아웃으로 그린다."""
    canvas = base_bgr.copy()
    height, width = canvas.shape[:2]
    scale = max(min(width, height) / 900.0, 0.55)
    font = cv2.FONT_HERSHEY_SIMPLEX

    for point in points:
        colour = PLUS if point.correction > 0 else MINUS
        centre = (int(point.x_px), int(point.y_px))
        cv2.circle(canvas, centre, max(int(5 * scale), 3), colour, -1)
        cv2.circle(canvas, centre, max(int(5 * scale), 3), (255, 255, 255),
                   max(int(1.4 * scale), 1))

        text = f"{'+' if point.correction > 0 else ''}{point.correction:.1f}"
        size, _ = cv2.getTextSize(text, font, 0.5 * scale, max(int(1.4 * scale), 1))
        # 콜아웃은 점 오른쪽 위. 화면 밖으로 나가면 반대편에 붙인다.
        box_x = centre[0] + int(11 * scale)
        box_y = centre[1] - int(11 * scale)
        if box_x + size[0] + 10 > width:
            box_x = centre[0] - int(11 * scale) - size[0] - 10
        if box_y - size[1] - 8 < 0:
            box_y = centre[1] + int(11 * scale) + size[1]

        pad = int(4 * scale)
        cv2.rectangle(canvas,
                      (box_x - pad, box_y - size[1] - pad),
                      (box_x + size[0] + pad, box_y + pad),
                      colour, -1)
        cv2.line(canvas, centre, (box_x, box_y - size[1] // 2), colour,
                 max(int(1.2 * scale), 1))
        cv2.putText(canvas, text, (box_x, box_y), font, 0.5 * scale,
                    (255, 255, 255), max(int(1.3 * scale), 1), cv2.LINE_AA)
    return canvas


def build_workbook(
    sheet_image_bgr: np.ndarray,
    points: list,
    part_no: str = "",
    part_name: str = "",
    process: str = "",
    material: str = "",
    control_no: str = "",
    applied_at: str | None = None,
    coefficient: float = 1.0,
) -> bytes:
    """현업 양식으로 채운 엑셀 파일을 바이트로 준다."""
    import openpyxl
    from openpyxl.drawing.image import Image as XlImage
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    if not TEMPLATE.exists():
        raise FileNotFoundError(f"보정시트 양식이 없습니다: {TEMPLATE}")

    book = openpyxl.load_workbook(TEMPLATE)
    page = book["00"]

    values = {
        "control_no": control_no or f"{part_no}-01",
        "part_name": part_name,
        "process": process or "OP10",
        "part_no": part_no,
        "material": material,
        "applied_at": applied_at or date.today().isoformat(),
    }
    for key, cell in HEADER_CELLS.items():
        page[cell] = values[key]

    # ── 그림 붙이기 ──────────────────────────────────────────
    # 양식 행 높이 14pt 기준으로 들어갈 자리를 계산한다(1pt = 4/3 px).
    box_height = int(IMAGE_ROWS * ROW_POINTS * 4 / 3)
    height, width = sheet_image_bgr.shape[:2]
    box_width = int(box_height * width / height)

    ok, buffer = cv2.imencode(".png", sheet_image_bgr)
    if not ok:
        raise ValueError("시트 그림을 PNG 로 만들지 못했습니다.")
    picture = XlImage(io.BytesIO(buffer.tobytes()))
    picture.width, picture.height = box_width, box_height
    page.add_image(picture, IMAGE_ANCHOR)

    # ── 포인트 표 ────────────────────────────────────────────
    table = book.create_sheet("포인트")
    headers = ["포인트", "X(px)", "Y(px)", "편차(mm)", "보정량(mm)", "방향"]
    table.append(headers)
    for index, name in enumerate(headers, start=1):
        cell = table.cell(1, index)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    for point in points:
        table.append([
            point.point_id, point.x_px, point.y_px,
            round(point.deviation, 2), round(point.correction, 2),
            "가공(살빼기)" if point.correction < 0 else "용접(살붙이기)",
        ])
    for index, size in enumerate((12, 10, 10, 12, 13, 16), start=1):
        table.column_dimensions[get_column_letter(index)].width = size
    table.freeze_panes = "A2"

    note = table.cell(len(points) + 3, 1)
    note.value = (f"보정 계수 {coefficient:.2f}x · 보정량은 작업자 수정을 "
                  f"반영한 최종값입니다.")
    note.font = Font(italic=True, size=9)

    stream = io.BytesIO()
    book.save(stream)
    return stream.getvalue()


__all__ = ["SheetPoint", "TEMPLATE", "draw_sheet_image", "build_workbook"]
